"""
Tests for questions app.
"""
from django.test import TestCase
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from apps.users.models import User, Role, UserRole
from .models import Question, Quiz, QuizQuestion


class QuestionAPITestCase(TestCase):
    """测试题目API"""
    
    def setUp(self):
        """设置测试数据"""
        # 创建角色
        self.admin_role = Role.objects.create(name='管理员', code='ADMIN')
        self.mentor_role = Role.objects.create(name='导师', code='MENTOR')
        self.student_role = Role.objects.create(name='学员', code='STUDENT')
        
        # 创建用户
        self.admin = User.objects.create_user(
            username='admin',
            password='admin123',
            real_name='管理员',
            employee_id='A001'
        )
        UserRole.objects.create(user=self.admin, role=self.admin_role)
        
        self.mentor = User.objects.create_user(
            username='mentor',
            password='mentor123',
            real_name='导师',
            employee_id='M001'
        )
        UserRole.objects.create(user=self.mentor, role=self.mentor_role)
        
        self.student = User.objects.create_user(
            username='student',
            password='student123',
            real_name='学员',
            employee_id='S001'
        )
        UserRole.objects.create(user=self.student, role=self.student_role)
        
        # 创建测试题目
        self.question = Question.objects.create(
            type='SINGLE',
            content='Python是什么类型的语言？',
            options={'A': '编译型', 'B': '解释型', 'C': '汇编', 'D': '机器语言'},
            correct_answer={'answer': 'B'},
            analysis='Python是解释型语言',
            difficulty=2,
            is_public=True,
            created_by=self.mentor
        )
        
        self.client = APIClient()
    
    def test_create_question_as_mentor(self):
        """测试导师创建题目"""
        self.client.force_authenticate(user=self.mentor)
        
        data = {
            'type': 'MULTIPLE',
            'content': '以下哪些是Python的特点？',
            'options': {'A': '简单易学', 'B': '开源', 'C': '跨平台', 'D': '编译型'},
            'correct_answer': {'answer': ['A', 'B', 'C']},
            'analysis': 'Python是简单易学、开源、跨平台的解释型语言',
            'difficulty': 3,
            'is_public': False
        }
        
        response = self.client.post('/api/questions/', data, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['type'], 'MULTIPLE')
    
    def test_create_question_as_student_fails(self):
        """测试学员创建题目失败"""
        self.client.force_authenticate(user=self.student)
        
        data = {
            'type': 'SINGLE',
            'content': '测试题目',
            'options': {'A': '选项A', 'B': '选项B'},
            'correct_answer': {'answer': 'A'},
            'difficulty': 1
        }
        
        response = self.client.post('/api/questions/', data, format='json')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_list_questions(self):
        """测试获取题目列表"""
        self.client.force_authenticate(user=self.student)
        
        response = self.client.get('/api/questions/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
    
    def test_get_question_detail(self):
        """测试获取题目详情"""
        self.client.force_authenticate(user=self.student)
        
        response = self.client.get(f'/api/questions/{self.question.id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['id'], self.question.id)
    
    def test_update_question_as_mentor(self):
        """测试导师更新题目"""
        self.client.force_authenticate(user=self.mentor)
        
        data = {
            'difficulty': 3
        }
        
        response = self.client.patch(f'/api/questions/{self.question.id}/', data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['difficulty'], 3)
    
    def test_delete_question(self):
        """测试删除题目"""
        self.client.force_authenticate(user=self.mentor)
        
        response = self.client.delete(f'/api/questions/{self.question.id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # 验证软删除
        self.question.refresh_from_db()
        self.assertTrue(self.question.is_deleted)


class QuizAPITestCase(TestCase):
    """测试测验API"""
    
    def setUp(self):
        """设置测试数据"""
        # 创建角色
        self.admin_role = Role.objects.create(name='管理员', code='ADMIN')
        self.mentor_role = Role.objects.create(name='导师', code='MENTOR')
        self.student_role = Role.objects.create(name='学员', code='STUDENT')
        
        # 创建用户
        self.mentor = User.objects.create_user(
            username='mentor',
            password='mentor123',
            real_name='导师',
            employee_id='M001'
        )
        UserRole.objects.create(user=self.mentor, role=self.mentor_role)
        
        self.student = User.objects.create_user(
            username='student',
            password='student123',
            real_name='学员',
            employee_id='S001'
        )
        UserRole.objects.create(user=self.student, role=self.student_role)
        
        # 创建测试题目
        self.question1 = Question.objects.create(
            type='SINGLE',
            content='题目1',
            options={'A': '选项A', 'B': '选项B'},
            correct_answer={'answer': 'A'},
            difficulty=2,
            is_public=True,
            created_by=self.mentor
        )
        
        self.question2 = Question.objects.create(
            type='MULTIPLE',
            content='题目2',
            options={'A': '选项A', 'B': '选项B', 'C': '选项C'},
            correct_answer={'answer': ['A', 'B']},
            difficulty=3,
            is_public=True,
            created_by=self.mentor
        )
        
        # 创建测试测验
        self.quiz = Quiz.objects.create(
            title='Python基础测验',
            description='测试Python基础知识',
            total_score=100,
            pass_score=60,
            is_public=True,
            created_by=self.mentor
        )
        
        self.client = APIClient()
    
    def test_create_quiz_as_mentor(self):
        """测试导师创建测验"""
        self.client.force_authenticate(user=self.mentor)
        
        data = {
            'title': '新测验',
            'description': '测试描述',
            'total_score': 100,
            'pass_score': 60,
            'is_public': False
        }
        
        response = self.client.post('/api/quizzes/', data, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['title'], '新测验')
    
    def test_list_quizzes(self):
        """测试获取测验列表"""
        self.client.force_authenticate(user=self.student)
        
        response = self.client.get('/api/quizzes/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
    
    def test_get_quiz_detail(self):
        """测试获取测验详情"""
        self.client.force_authenticate(user=self.student)
        
        response = self.client.get(f'/api/quizzes/{self.quiz.id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['id'], self.quiz.id)
    
    def test_add_questions_to_quiz(self):
        """测试添加题目到测验"""
        self.client.force_authenticate(user=self.mentor)
        
        data = {
            'questions': [
                {
                    'question_id': self.question1.id,
                    'score': 50,
                    'sort_order': 1
                },
                {
                    'question_id': self.question2.id,
                    'score': 50,
                    'sort_order': 2
                }
            ]
        }
        
        response = self.client.post(f'/api/quizzes/{self.quiz.id}/add-questions/', data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # 验证题目已添加
        self.assertEqual(self.quiz.quiz_questions.count(), 2)
    
    def test_remove_question_from_quiz(self):
        """测试从测验中移除题目"""
        self.client.force_authenticate(user=self.mentor)
        
        # 先添加题目
        QuizQuestion.objects.create(
            quiz=self.quiz,
            question=self.question1,
            sort_order=1,
            score=50
        )
        
        # 移除题目
        response = self.client.delete(
            f'/api/quizzes/{self.quiz.id}/remove-question/?question_id={self.question1.id}'
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # 验证题目已移除
        self.assertEqual(self.quiz.quiz_questions.count(), 0)
    
    def test_reorder_questions(self):
        """测试重新排序题目"""
        self.client.force_authenticate(user=self.mentor)
        
        # 先添加题目
        QuizQuestion.objects.create(
            quiz=self.quiz,
            question=self.question1,
            sort_order=1,
            score=50
        )
        QuizQuestion.objects.create(
            quiz=self.quiz,
            question=self.question2,
            sort_order=2,
            score=50
        )
        
        # 重新排序
        data = {
            'question_orders': [
                {'question_id': self.question1.id, 'sort_order': 2},
                {'question_id': self.question2.id, 'sort_order': 1}
            ]
        }
        
        response = self.client.put(f'/api/quizzes/{self.quiz.id}/reorder-questions/', data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # 验证顺序已更新
        qq1 = QuizQuestion.objects.get(quiz=self.quiz, question=self.question1)
        qq2 = QuizQuestion.objects.get(quiz=self.quiz, question=self.question2)
        self.assertEqual(qq1.sort_order, 2)
        self.assertEqual(qq2.sort_order, 1)
    
    def test_get_quiz_questions(self):
        """测试获取测验题目列表"""
        self.client.force_authenticate(user=self.student)
        
        # 先添加题目
        QuizQuestion.objects.create(
            quiz=self.quiz,
            question=self.question1,
            sort_order=1,
            score=50
        )
        QuizQuestion.objects.create(
            quiz=self.quiz,
            question=self.question2,
            sort_order=2,
            score=50
        )
        
        response = self.client.get(f'/api/quizzes/{self.quiz.id}/questions/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 2)



class QuestionImportTestCase(TestCase):
    """测试题目导入功能"""
    
    def setUp(self):
        """设置测试数据"""
        # 创建角色
        self.admin_role = Role.objects.create(name='管理员', code='ADMIN')
        self.mentor_role = Role.objects.create(name='导师', code='MENTOR')
        self.student_role = Role.objects.create(name='学员', code='STUDENT')
        
        # 创建用户
        self.mentor = User.objects.create_user(
            username='mentor',
            password='mentor123',
            real_name='导师',
            employee_id='M001'
        )
        UserRole.objects.create(user=self.mentor, role=self.mentor_role)
        
        self.student = User.objects.create_user(
            username='student',
            password='student123',
            real_name='学员',
            employee_id='S001'
        )
        UserRole.objects.create(user=self.student, role=self.student_role)
        
        self.client = APIClient()
    
    def test_import_questions_success(self):
        """测试成功导入题目"""
        import os
        import json
        from openpyxl import Workbook
        import tempfile
        
        # 创建测试Excel文件
        wb = Workbook()
        ws = wb.active
        
        # 添加表头
        ws.append(['题目类型', '题目内容', '选项', '正确答案', '题目解析', '难度', '是否公开'])
        
        # 添加测试数据
        ws.append([
            'SINGLE',
            'Python是什么类型的语言？',
            json.dumps({"A": "编译型", "B": "解释型"}, ensure_ascii=False),
            json.dumps({"answer": "B"}, ensure_ascii=False),
            '解析内容',
            2,
            True
        ])
        
        # 保存到临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # 执行导入
            self.client.force_authenticate(user=self.mentor)
            
            with open(tmp_path, 'rb') as f:
                response = self.client.post(
                    '/api/questions/import/',
                    {'file': f},
                    format='multipart'
                )
            
            self.assertEqual(response.status_code, status.HTTP_200_OK)
            self.assertTrue(response.data['success'])
            self.assertEqual(response.data['data']['success_count'], 1)
            self.assertEqual(response.data['data']['error_count'], 0)
            
            # 验证题目已创建
            self.assertEqual(Question.objects.count(), 1)
            question = Question.objects.first()
            self.assertEqual(question.type, 'SINGLE')
            self.assertEqual(question.created_by, self.mentor)
        
        finally:
            # 清理临时文件
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_import_questions_with_errors(self):
        """测试导入包含错误的题目"""
        import os
        import json
        from openpyxl import Workbook
        import tempfile
        
        # 创建测试Excel文件
        wb = Workbook()
        ws = wb.active
        
        # 添加表头
        ws.append(['题目类型', '题目内容', '选项', '正确答案', '题目解析', '难度', '是否公开'])
        
        # 添加正确的数据
        ws.append([
            'SINGLE',
            '正确的题目',
            json.dumps({"A": "选项A", "B": "选项B"}, ensure_ascii=False),
            json.dumps({"answer": "A"}, ensure_ascii=False),
            '解析',
            2,
            True
        ])
        
        # 添加错误的数据（缺少题目内容）
        ws.append([
            'SINGLE',
            '',  # 空内容
            json.dumps({"A": "选项A"}, ensure_ascii=False),
            json.dumps({"answer": "A"}, ensure_ascii=False),
            '解析',
            2,
            True
        ])
        
        # 保存到临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # 执行导入
            self.client.force_authenticate(user=self.mentor)
            
            with open(tmp_path, 'rb') as f:
                response = self.client.post(
                    '/api/questions/import/',
                    {'file': f},
                    format='multipart'
                )
            
            self.assertEqual(response.status_code, status.HTTP_200_OK)
            self.assertTrue(response.data['success'])
            self.assertEqual(response.data['data']['success_count'], 1)
            self.assertEqual(response.data['data']['error_count'], 1)
            
            # 验证只有正确的题目被创建
            self.assertEqual(Question.objects.count(), 1)
        
        finally:
            # 清理临时文件
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_import_questions_as_student_fails(self):
        """测试学员无法导入题目"""
        import json
        from openpyxl import Workbook
        import tempfile
        
        # 创建测试Excel文件
        wb = Workbook()
        ws = wb.active
        ws.append(['题目类型', '题目内容', '选项', '正确答案', '题目解析', '难度', '是否公开'])
        ws.append([
            'SINGLE',
            '测试题目',
            json.dumps({"A": "选项A"}, ensure_ascii=False),
            json.dumps({"answer": "A"}, ensure_ascii=False),
            '解析',
            2,
            True
        ])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # 学员尝试导入
            self.client.force_authenticate(user=self.student)
            
            with open(tmp_path, 'rb') as f:
                response = self.client.post(
                    '/api/questions/import/',
                    {'file': f},
                    format='multipart'
                )
            
            # 应该返回403禁止访问
            self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        
        finally:
            # 清理临时文件
            import os
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_import_invalid_file_format(self):
        """测试导入无效的文件格式"""
        import tempfile
        
        # 创建一个非Excel文件
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp.write(b'This is not an Excel file')
            tmp_path = tmp.name
        
        try:
            self.client.force_authenticate(user=self.mentor)
            
            with open(tmp_path, 'rb') as f:
                response = self.client.post(
                    '/api/questions/import/',
                    {'file': f},
                    format='multipart'
                )
            
            # 应该返回400错误
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
            self.assertFalse(response.data['success'])
        
        finally:
            # 清理临时文件
            import os
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
